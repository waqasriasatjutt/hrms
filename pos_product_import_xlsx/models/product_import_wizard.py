import base64
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


class PosProductImportWizard(models.TransientModel):
    _name = 'pos.product.import.wizard'
    _description = 'Import POS Products from XLSX'

    file_data = fields.Binary(string='XLSX File', required=True, attachment=False)
    file_name = fields.Char(string='File Name')
    result_message = fields.Text(string='Result', readonly=True)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Done'),
    ], default='draft')

    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_(
                'Python library "openpyxl" is not installed.\n'
                'Run: pip install openpyxl'
            ))
        if not self.file_data:
            raise UserError(_('Please upload an XLSX file.'))

        file_name = self.file_name or ''
        if not file_name.lower().endswith('.xlsx'):
            raise UserError(_('Only .xlsx files are supported.'))

        raw = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active

        # Read header row
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        required_cols = ['Barcode', 'Sale Price']
        for col in required_cols:
            if col not in headers:
                raise UserError(_(
                    'Missing required column: "%s"\n\nFound columns: %s'
                ) % (col, ', '.join(headers)))

        col_index = {name: idx for idx, name in enumerate(headers)}

        created = 0
        updated = 0
        skipped = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            def get(col_name):
                idx = col_index.get(col_name)
                if idx is None:
                    return None
                val = row[idx]
                if val is None:
                    return None
                return str(val).strip()

            barcode = get('Barcode')
            if not barcode:
                skipped += 1
                continue

            try:
                vals = self._build_product_vals(row_num, get)
                product = self._find_product(barcode)
                if product:
                    product.write(vals)
                    updated += 1
                else:
                    if not vals.get('name'):
                        vals['name'] = barcode
                    self.env['product.template'].create(vals)
                    created += 1
            except Exception as e:
                errors.append('Row %d (barcode=%s): %s' % (row_num, barcode, str(e)))
                _logger.warning('Import error row %d: %s', row_num, e)

        lines = [
            'Import complete.',
            'Created : %d' % created,
            'Updated : %d' % updated,
            'Skipped : %d (no barcode)' % skipped,
        ]
        if errors:
            lines.append('\nErrors (%d):' % len(errors))
            lines.extend(errors)

        self.write({
            'result_message': '\n'.join(lines),
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'pos.product.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _find_product(self, barcode):
        """Return product.template or False."""
        tmpl = self.env['product.template'].search(
            [('barcode', '=', barcode)], limit=1
        )
        if tmpl:
            return tmpl
        # Also check product.product (variant barcode)
        product = self.env['product.product'].search(
            [('barcode', '=', barcode)], limit=1
        )
        return product.product_tmpl_id if product else False

    def _build_product_vals(self, row_num, get):
        vals = {}

        barcode = get('Barcode')
        if barcode:
            vals['barcode'] = barcode

        # Sale Price
        sale_price = get('Sale Price')
        if sale_price:
            try:
                vals['list_price'] = float(sale_price)
            except ValueError:
                raise UserError(_('Row %d: Invalid Sale Price "%s"') % (row_num, sale_price))

        # Cost
        cost = get('Cost')
        if cost:
            try:
                vals['standard_price'] = float(cost)
            except ValueError:
                raise UserError(_('Row %d: Invalid Cost "%s"') % (row_num, cost))

        # Available in POS
        avail_pos = get('Available in POS')
        if avail_pos is not None:
            vals['available_in_pos'] = avail_pos.lower() in ('1', 'true', 'yes', 'y')

        # Product Category
        categ_name = get('Product Category')
        if categ_name:
            categ = self.env['product.category'].search(
                [('name', '=ilike', categ_name)], limit=1
            )
            if categ:
                vals['categ_id'] = categ.id
            else:
                new_categ = self.env['product.category'].create({'name': categ_name})
                vals['categ_id'] = new_categ.id

        # POS Category
        # POS Category
        # pos_categ_name = get('Point of Sale Category')
        # if pos_categ_name:
        #     pos_categ = self.env['pos.category'].search(
        #         [('name', '=ilike', pos_categ_name)], limit=1
        #     )
        #     if pos_categ:
        #         vals['pos_category_id'] = pos_categ.id
        #     else:
        #         new_pos_categ = self.env['pos.category'].create({'name': pos_categ_name})
        #         vals['pos_category_id'] = new_pos_categ.id

        # POS Category
        pos_categ_name = get('Point of Sale Category')
        if pos_categ_name:
            pos_categ = self.env['pos.category'].search(
                [('name', '=ilike', pos_categ_name)], limit=1
            )
            if pos_categ:
                vals['pos_categ_ids'] = pos_categ.id
            else:
                new_pos_categ = self.env['pos.category'].create({'name': pos_categ_name})
                vals['pos_categ_ids'] = new_pos_categ.id

        
        
        # pos_categ_name = get('Point of Sale Category')
        # if pos_categ_name:
        #     pos_categ = self.env['pos.category'].search(
        #         [('name', '=ilike', pos_categ_name)], limit=1
        #     )
        #     if not pos_categ:
        #         pos_categ = self.env['pos.category'].create({'name': pos_categ_name})
        
        #     vals['pos_categ_ids'] = [(6, 0, [pos_categ.id])]

        # Supplier
        supplier_name = get('Supplier')
        if supplier_name:
            partner = self.env['res.partner'].search(
                [('name', '=ilike', supplier_name), ('supplier_rank', '>', 0)],
                limit=1,
            )
            if not partner:
                partner = self.env['res.partner'].search(
                    [('name', '=ilike', supplier_name)], limit=1
                )
            if partner:
                vals['seller_ids'] = [(0, 0, {
                    'partner_id': partner.id,
                    'price': vals.get('standard_price', 0.0),
                })]

        # Sale Taxes
        sale_tax_name = get('Sale Taxes')
        if sale_tax_name:
            taxes = self._resolve_taxes(sale_tax_name, 'sale')
            if taxes:
                vals['taxes_id'] = [(6, 0, taxes.ids)]

        # Purchase Taxes
        purchase_tax_name = get('Purchase Taxes')
        if purchase_tax_name:
            taxes = self._resolve_taxes(purchase_tax_name, 'purchase')
            if taxes:
                vals['supplier_taxes_id'] = [(6, 0, taxes.ids)]

        return vals

    def _resolve_taxes(self, tax_str, type_tax_use):
        """Support comma-separated tax names or single name."""
        Tax = self.env['account.tax']
        names = [n.strip() for n in tax_str.split(',') if n.strip()]
        found = Tax
        for name in names:
            tax = Tax.search([
                ('name', '=ilike', name),
                ('type_tax_use', '=', type_tax_use),
            ], limit=1)
            if tax:
                found |= tax
        return found
